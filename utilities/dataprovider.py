import openpyxl


def get_data(sheetName):

    workbook=openpyxl.load_workbook("..//excel//testdata.xlsx")
    sheet=workbook[sheetName]
    totalrows=sheet.max_row
    totalcolumns=sheet.max_column
    mainlist=[]
    for i in range(2,totalrows+1):
        dataList=[]
        for j in range(1,totalcolumns+1):
            data=sheet.cell(row=i,column=j).value
            dataList.append(data)
        mainlist.insert(i,dataList)
    return mainlist